#!/usr/bin/env python3
"""Write boarding check-in/out rows (JSON array) to Excel. Used when SQL is run via Supabase MCP."""

import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

def format_check_in_status(code: str) -> str:
    if not code:
        return ""
    return code.replace("_", " ").title()


HEADERS = [
    "Booking ref",
    "Check-in status",
    "Status code",
    "Check-in date",
    "Check-out date",
    "Actual check-in",
    "Actual check-out",
    "Client name",
    "Client ID",
    "Pet name",
    "Pet ID",
    "Booking ID",
]


def format_ts(value: Optional[str]) -> str:
    if not value:
        return ""
    try:
        dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d %H:%M")
    except ValueError:
        return value


def main() -> None:
    if len(sys.argv) < 3:
        print("Usage: json_to_boarding_xlsx.py <input.json> <output.xlsx>", file=sys.stderr)
        sys.exit(1)

    in_path = Path(sys.argv[1])
    out_path = Path(sys.argv[2])
    raw = in_path.read_text(encoding="utf-8")

    payload = raw
    if raw.lstrip().startswith("{"):
        outer = json.loads(raw)
        payload = outer.get("result", raw)

    # Supabase MCP: array inside <untrusted-data-...> … </untrusted-data-...>
    if "<untrusted-data-" in payload:
        open_tag = payload.find("<untrusted-data-")
        close_tag = payload.find(">", open_tag)
        close_block = payload.find("</untrusted-data-", close_tag)
        if close_tag != -1 and close_block != -1:
            payload = payload[close_tag + 1 : close_block].strip()

    start = payload.find("[")
    end = payload.rfind("]")
    if start == -1 or end == -1:
        raise SystemExit("No JSON array found in input")
    rows = json.loads(payload[start : end + 1])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Boarding"

    ws.append(HEADERS)

    for row in rows:
        status_code = row.get("status") or ""
        ws.append(
            [
                row.get("booking_ref") or "",
                format_check_in_status(status_code),
                status_code,
                row.get("check_in_date") or "",
                row.get("check_out_date") or "",
                format_ts(row.get("actual_check_in_at")),
                format_ts(row.get("actual_check_out_at")),
                row.get("client_name") or "",
                row.get("client_id") or "",
                row.get("pet_name") or "",
                row.get("pet_id") or "",
                row.get("booking_id") or "",
            ]
        )

    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {len(rows)} rows → {out_path}")


if __name__ == "__main__":
    main()
